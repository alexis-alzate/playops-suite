import getpass
import socket
from datetime import datetime

from app_procesos.config import APP_TITLE, APP_VERSION, REPORTS_DIR, USER_CONFIG_FILE
from app_procesos.shared.activity import obtener_historial_dashboard


def exportar_resumen_ejecutivo():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as error:
        raise RuntimeError("Falta instalar openpyxl para exportar el resumen ejecutivo.") from error

    output_dir = REPORTS_DIR / "resumenes_ejecutivos"
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now()
    path = output_dir / f"resumen_ejecutivo_{timestamp.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen ejecutivo"

    dark = "0E1A23"
    panel = "132B39"
    accent = "14A8CC"
    light = "EAF5FA"
    white = "FFFFFF"
    border = Border(
        left=Side(style="thin", color="9FB8C5"),
        right=Side(style="thin", color="9FB8C5"),
        top=Side(style="thin", color="9FB8C5"),
        bottom=Side(style="thin", color="9FB8C5"),
    )

    sheet.merge_cells("A1:F1")
    sheet["A1"] = "PlayOps Suite"
    sheet["A1"].fill = PatternFill("solid", fgColor=dark)
    sheet["A1"].font = Font(color=white, bold=True, size=18)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells("A2:F2")
    sheet["A2"] = "Apostamos por la tecnologia"
    sheet["A2"].fill = PatternFill("solid", fgColor=panel)
    sheet["A2"].font = Font(color=light, italic=True, size=11)
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    metadata = [
        ("Aplicacion", APP_TITLE),
        ("Version", APP_VERSION),
        ("Responsable Windows", getpass.getuser()),
        ("Equipo", socket.gethostname()),
        ("Fecha", timestamp.strftime("%Y-%m-%d")),
        ("Hora", timestamp.strftime("%H:%M:%S")),
        ("Configuracion", str(USER_CONFIG_FILE)),
    ]
    row_index = 4
    for label, value in metadata:
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 1).font = Font(bold=True, color=dark)
        sheet.cell(row_index, 2).font = Font(color=dark)
        sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=light)
        sheet.cell(row_index, 2).fill = PatternFill("solid", fgColor="F7FBFD")
        sheet.cell(row_index, 1).border = border
        sheet.cell(row_index, 2).border = border
        row_index += 1

    row_index += 2
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=6)
    sheet.cell(row_index, 1, "Historial operativo reciente")
    sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=accent)
    sheet.cell(row_index, 1).font = Font(color=white, bold=True, size=12)
    sheet.cell(row_index, 1).alignment = Alignment(horizontal="center")
    row_index += 1

    headers = ["Proceso", "Estado", "Detalle", "Ruta"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, column, header)
        cell.fill = PatternFill("solid", fgColor=panel)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    row_index += 1

    for item in obtener_historial_dashboard():
        values = [
            item.get("proceso", ""),
            item.get("estado", ""),
            item.get("detalle", ""),
            item.get("ruta", ""),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.fill = PatternFill("solid", fgColor="F7FBFD")
            cell.font = Font(color=dark)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row_index += 1

    widths = [24, 18, 55, 75, 18, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A12"
    workbook.save(path)
    return path
