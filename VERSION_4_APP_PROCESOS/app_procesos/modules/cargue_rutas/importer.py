import csv
from pathlib import Path

from .models import ArchivoClientePreview
from .validator import validar_archivo


MAX_SAMPLE_ROWS = 12


def cargar_preview_cliente(path, *, max_rows=MAX_SAMPLE_ROWS, header_row=1, data_start_row=None):
    archivo = validar_archivo(path, permite_macro=True)
    header_row = _validar_numero_fila(header_row, "fila de encabezados")
    data_start_row = _validar_numero_fila(data_start_row or header_row + 1, "fila inicial de clientes")
    if data_start_row <= header_row:
        raise ValueError("La fila inicial de clientes debe ser mayor que la fila de encabezados.")

    suffix = archivo.suffix.lower()
    if suffix == ".csv":
        return _leer_csv(archivo, max_rows=max_rows, header_row=header_row, data_start_row=data_start_row)
    if suffix in {".xlsx", ".xlsm"}:
        return _leer_excel_openpyxl(archivo, max_rows=max_rows, header_row=header_row, data_start_row=data_start_row)
    if suffix == ".xls":
        raise ValueError(
            "Los archivos .xls antiguos no se pueden leer todavia en este modulo. "
            "Si es la plantilla soporte, carguelo en el panel derecho. "
            "Si es el archivo del cliente, guardelo como .xlsx y vuelva a cargarlo."
        )
    raise ValueError(f"Formato no soportado: {suffix}")


def _leer_csv(path, *, max_rows, header_row, data_start_row):
    encoding = _detectar_encoding(path)
    with path.open("r", encoding=encoding, newline="") as file:
        sample = file.read(4096)
        file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = list(csv.reader(file, dialect=dialect))
        if not reader:
            raise ValueError("El archivo del cliente esta vacio.")
        if header_row > len(reader):
            raise ValueError("La fila de encabezados no existe en el archivo.")
        header = reader[header_row - 1]
        columnas = _limpiar_columnas(header)
        filas = []
        for row_number, values in enumerate(reader[data_start_row - 1 :], start=data_start_row):
            if _es_fila_instruccion(values):
                continue
            fila = {}
            for index, columna in enumerate(columnas):
                value = values[index] if index < len(values) else ""
                fila[columna] = _to_text(value)
            if any(fila.values()):
                fila["__fila_origen"] = str(row_number)
                filas.append(fila)
            if max_rows is not None and len(filas) >= max_rows:
                break
    return ArchivoClientePreview(path, "CSV", columnas, filas, len(filas))


def _leer_excel_openpyxl(path, *, max_rows, header_row, data_start_row):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Falta instalar openpyxl para leer Excel. Ejecute: python -m pip install openpyxl"
        ) from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    if sheet.max_row < header_row:
        workbook.close()
        raise ValueError("La fila de encabezados no existe en el archivo.")

    header_cells = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=False))
    columnas = _limpiar_columnas([cell.value for cell in header_cells])
    filas = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=False), start=data_start_row):
        values = [cell.value for cell in cells]
        if _es_fila_instruccion(values):
            continue
        fila = {}
        for index, columna in enumerate(columnas):
            cell = cells[index] if index < len(cells) else None
            fila[columna] = _to_text(cell.value if cell else "", getattr(cell, "number_format", ""))
        if any(fila.values()):
            fila["__fila_origen"] = str(row_number)
            filas.append(fila)
        if max_rows is not None and len(filas) >= max_rows:
            break
    workbook.close()
    return ArchivoClientePreview(path, sheet.title, columnas, filas, len(filas))


def _limpiar_columnas(columnas):
    resultado = []
    usados = {}
    for index, columna in enumerate(columnas, start=1):
        nombre = _to_text(columna).strip() or f"Columna {index}"
        base = nombre
        usados[base] = usados.get(base, 0) + 1
        if usados[base] > 1:
            nombre = f"{base} ({usados[base]})"
        resultado.append(nombre)
    return resultado


def _validar_numero_fila(value, label):
    try:
        number = int(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"La {label} debe ser un numero entero.") from error
    if number < 1:
        raise ValueError(f"La {label} debe ser mayor o igual a 1.")
    return number


def _to_text(value, number_format=""):
    if value is None:
        return ""
    if isinstance(value, (int, float)) and "%" in str(number_format or ""):
        return f"{float(value) * 100}%"
    return str(value).strip()


def _es_fila_instruccion(values):
    texto = " ".join(_to_text(value).lower() for value in values if value is not None)
    if not texto:
        return False
    marcadores = [
        "obligatorio",
        "opcional",
        "ejemplo",
        "indique",
        "solo si se tiene",
    ]
    return sum(1 for marcador in marcadores if marcador in texto) >= 2


def _detectar_encoding(path):
    data = Path(path).read_bytes()[:4]
    if data.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    return "utf-8"
