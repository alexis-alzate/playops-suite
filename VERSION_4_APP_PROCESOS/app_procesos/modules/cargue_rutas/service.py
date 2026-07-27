from datetime import datetime
from pathlib import Path
import subprocess

from app_procesos.config import REPORTS_DIR

from .calculator import calcular
from .cleaner import limpiar_datos
from .corrections import aplicar_correcciones
from .duplicate_review import aplicar_revision_duplicados
from .exporter import exportar_resultado
from .importer import cargar_preview_cliente
from .listo_csv_writer import generar_csv_listo, vista_previa_csv_listo
from .mapper import sugerir_mapeo, puntaje_columnas_requeridas
from .normalizer import normalizar_registros
from .reconciler import aplicar_formula_plantilla
from .template_writer import generar_archivo_final
from .validator import validar_archivo

ONEDRIVE_EMPRESA = "OneDrive - PLAY TECHNOLOGIES S.A.S"
RUTAS_AUDITORIA_ONEDRIVE = [
    Path("DOCUMENTACION LISTO") / "Procesos 7 AM" / "Monitoreo 7AM",
    Path("Área de Soporte - Monitoreo 7AM"),
]
AUDITORIA_CENTRAL_DIR = "Control Cargues"
AUDITORIA_CENTRAL_FILE = "bitacora_cargues.xlsx"
ULTIMO_ESTADO_ONEDRIVE = ""
ULTIMO_RESUMEN_EJECUTIVO = ""


def analizar_archivo_cliente(archivo_cliente, *, header_row=1, data_start_row=3, auto_detect=False):
    if auto_detect:
        detectado = detectar_estructura_archivo_cliente(archivo_cliente)
        if detectado:
            header_row = detectado["header_row"]
            data_start_row = detectado["data_start_row"]

    preview = cargar_preview_cliente(
        archivo_cliente,
        header_row=header_row,
        data_start_row=data_start_row,
    )
    mapeo = sugerir_mapeo(preview.columnas)
    return preview, mapeo


def detectar_estructura_archivo_cliente(archivo_cliente, *, max_header_row=12):
    mejor = None
    for header_row in range(1, max_header_row + 1):
        for offset in (1, 2, 3):
            data_start_row = header_row + offset
            try:
                preview = cargar_preview_cliente(
                    archivo_cliente,
                    max_rows=8,
                    header_row=header_row,
                    data_start_row=data_start_row,
                )
            except Exception:
                continue
            if not preview.columnas or not preview.filas_muestra:
                continue

            requeridos, puntaje = puntaje_columnas_requeridas(preview.columnas)
            columnas_utiles = sum(1 for columna in preview.columnas if not str(columna).lower().startswith("columna "))
            filas_con_datos = len(preview.filas_muestra)
            score = (requeridos * 1000) + puntaje + (columnas_utiles * 5) + filas_con_datos
            candidato = {
                "header_row": header_row,
                "data_start_row": data_start_row,
                "score": score,
                "requeridos": requeridos,
                "puntaje": puntaje,
            }
            if (
                mejor is None
                or candidato["score"] > mejor["score"]
                or (
                    candidato["score"] == mejor["score"]
                    and _prioridad_fila_datos(candidato) < _prioridad_fila_datos(mejor)
                )
            ):
                mejor = candidato

    if mejor and mejor["requeridos"] >= 2:
        return mejor
    return None


def _prioridad_fila_datos(candidato):
    esperada = candidato["header_row"] + 2
    return (
        abs(candidato["data_start_row"] - esperada),
        candidato["data_start_row"],
    )


def preparar_cargue(
    archivo_cliente,
    *,
    header_row=1,
    data_start_row=3,
    mapping=None,
    interes_default=None,
    periodicidad_default_codigo=None,
    periodicidad_override_codigo=None,
):
    if mapping is None and int(header_row) == 1 and int(data_start_row) == 3:
        detectado = detectar_estructura_archivo_cliente(archivo_cliente)
        if detectado:
            header_row = detectado["header_row"]
            data_start_row = detectado["data_start_row"]

    preview = cargar_preview_cliente(
        archivo_cliente,
        max_rows=None,
        header_row=header_row,
        data_start_row=data_start_row,
    )
    mapeo = mapping or sugerir_mapeo(preview.columnas)
    normalizacion = normalizar_registros(
        preview,
        mapeo,
        interes_default=interes_default,
        periodicidad_default_codigo=periodicidad_default_codigo,
        periodicidad_override_codigo=periodicidad_override_codigo,
    )
    return preview, mapeo, normalizacion


def validar_cargue(
    archivo_cliente,
    *,
    header_row=1,
    data_start_row=3,
    mapping=None,
    interes_default=None,
    periodicidad_default_codigo=None,
    periodicidad_override_codigo=None,
    correcciones=None,
    decisiones_duplicados=None,
    responsable="",
    auditoria=None,
):
    auditoria = auditoria or _auditoria_desde_responsable(responsable)
    cliente = validar_archivo(archivo_cliente, permite_macro=True)
    preview, mapeo, normalizacion = preparar_cargue(
        cliente,
        header_row=header_row,
        data_start_row=data_start_row,
        mapping=mapping,
        interes_default=interes_default,
        periodicidad_default_codigo=periodicidad_default_codigo,
        periodicidad_override_codigo=periodicidad_override_codigo,
    )
    normalizacion = aplicar_correcciones(normalizacion, correcciones)
    normalizacion = aplicar_formula_plantilla(normalizacion)
    normalizacion, _duplicados = aplicar_revision_duplicados(
        normalizacion,
        decisiones=decisiones_duplicados,
    )
    return preview, mapeo, normalizacion


def procesar_archivos(
    archivo_cliente,
    archivo_soporte=None,
    *,
    header_row=1,
    data_start_row=3,
    mapping=None,
    interes_default=None,
    dias_credito_default=30,
    fecha_prox_pago_mode="vacia",
    fecha_prox_pago_manual="",
    periodicidad_default_codigo=None,
    periodicidad_override_codigo=None,
    correcciones=None,
    decisiones_duplicados=None,
    responsable="",
    auditoria=None,
):
    auditoria = auditoria or _auditoria_desde_responsable(responsable)
    cliente = validar_archivo(archivo_cliente, permite_macro=True)
    preview, mapeo, normalizacion = validar_cargue(
        cliente,
        header_row=header_row,
        data_start_row=data_start_row,
        mapping=mapping,
        interes_default=interes_default,
        periodicidad_default_codigo=periodicidad_default_codigo,
        periodicidad_override_codigo=periodicidad_override_codigo,
        correcciones=correcciones,
        decisiones_duplicados=decisiones_duplicados,
    )
    if normalizacion.total_con_error:
        errores_bloqueantes = [
            registro for registro in normalizacion.registros if registro.errores
        ][:10]
        detalle = "; ".join(
            f"fila {registro.fila_origen}: {', '.join(registro.errores)}"
            for registro in errores_bloqueantes
        )
        raise ValueError(
            "El cargue tiene errores que deben revisarse antes de generar el archivo final. "
            "Use 'Validar datos', corrija o confirme las diferencias y luego pulse 'Aplicar correcciones'. "
            + detalle
        )

    archivo_csv, archivo_csv_listo = generar_csv_listo(
        normalizacion,
        archivo_cliente=cliente,
        dias_credito_default=dias_credito_default,
        fecha_prox_pago_mode=fecha_prox_pago_mode,
        fecha_prox_pago_manual=fecha_prox_pago_manual,
    )
    try:
        archivo_final = generar_archivo_final(
            normalizacion,
            archivo_cliente=cliente,
            dias_credito_default=dias_credito_default,
            fecha_prox_pago_mode=fecha_prox_pago_mode,
            fecha_prox_pago_manual=fecha_prox_pago_manual,
        )
        advertencia_excel = ""
    except Exception as exc:
        archivo_final = "No generado"
        advertencia_excel = f"\nAdvertencia Excel respaldo: {exc}\n"
    bitacora = generar_bitacora_cargue(
        normalizacion,
        archivo_cliente=cliente,
        archivo_csv=archivo_csv,
        archivo_csv_listo=archivo_csv_listo or archivo_csv,
        archivo_excel=archivo_final,
        mapeo=mapeo,
        preview=preview,
        header_row=header_row,
        data_start_row=data_start_row,
        auditoria=auditoria,
        advertencia_excel=advertencia_excel,
    )
    campos_detectados = _formatear_mapeo(mapeo)
    resumen_periodicidad = _resumen_periodicidad(normalizacion.registros)
    resumen_abonos = _resumen_abonos(normalizacion.registros)
    registros_con_error = [
        registro for registro in normalizacion.registros if registro.errores
    ][:20]
    errores = [
        f"- Fila {registro.fila_origen} / doc {registro.documento or 'sin documento'}: "
        + "; ".join(registro.errores)
        for registro in registros_con_error
    ]
    datos = (
        "Archivos recibidos:\n"
        + _formatear_auditoria(auditoria)
        + "\n"
        f"- Cliente: {cliente}\n"
        "- Plantilla: interna de la aplicacion\n"
        f"- Archivo Excel respaldo: {archivo_final}\n"
        f"- Archivo CSV generado: {archivo_csv}\n"
        f"- Archivo para cargar al sistema: {archivo_csv_listo or archivo_csv}\n\n"
        f"- Bitacora del cargue: {bitacora}\n\n"
        f"{advertencia_excel}"
        f"Hoja detectada: {preview.hoja}\n"
        f"Fila encabezados: {header_row}\n"
        f"Fila inicial clientes: {data_start_row}\n"
        f"Columnas detectadas: {len(preview.columnas)}\n"
        f"Registros leidos: {normalizacion.total_registros}\n"
        f"Registros validos: {normalizacion.total_validos}\n"
        f"Registros con error: {normalizacion.total_con_error}\n"
        f"Con abono informado: {normalizacion.total_con_abono_informado}\n"
        f"Requieren reconstruccion de abono: {normalizacion.total_requiere_reconstruccion}\n\n"
        "Estado de abonos:\n"
        + "\n".join(resumen_abonos)
        + "\n\n"
        "Periodicidad detectada:\n"
        + "\n".join(resumen_periodicidad)
        + "\n\n"
        "Mapeo sugerido:\n"
        + "\n".join(campos_detectados)
        + "\n\nErrores generales:\n"
        + ("\n".join(normalizacion.errores_generales) if normalizacion.errores_generales else "- Sin errores generales")
        + "\n\nPrimeros errores por fila:\n"
        + ("\n".join(errores) if errores else "- Sin errores en la validacion base")
    )
    datos = limpiar_datos(datos)
    resultado = calcular(datos)
    salida = REPORTS_DIR / "cargue_rutas_resultado.txt"
    exportar_resultado(resultado, salida)
    return archivo_csv_listo or archivo_csv


def resumen_validacion(normalizacion):
    errores_bloqueantes = sum(1 for registro in normalizacion.registros if registro.errores)
    abonos_reconstruidos = sum(1 for registro in normalizacion.registros if registro.abono_estado == "reconstruido")
    abonos_pendientes = sum(1 for registro in normalizacion.registros if registro.abono_estado == "requiere_reconstruccion")
    ajustes_automaticos = [
        error for error in normalizacion.errores_generales if "Se ajustaron" in error
    ]
    advertencias = [
        error for error in normalizacion.errores_generales if error not in ajustes_automaticos
    ]
    return {
        "total": normalizacion.total_registros,
        "validos": normalizacion.total_validos,
        "errores_bloqueantes": errores_bloqueantes,
        "advertencias": len(advertencias),
        "ajustes_automaticos": len(ajustes_automaticos) + abonos_reconstruidos,
        "abonos_informados": normalizacion.total_con_abono_informado,
        "abonos_reconstruidos": abonos_reconstruidos,
        "abonos_pendientes": abonos_pendientes,
        "errores_generales": normalizacion.errores_generales,
    }


def preview_csv_final(
    normalizacion,
    *,
    dias_credito_default=30,
    fecha_prox_pago_mode="vacia",
    fecha_prox_pago_manual="",
    limite=8,
):
    headers, rows = vista_previa_csv_listo(
        normalizacion,
        dias_credito_default=dias_credito_default,
        fecha_prox_pago_mode=fecha_prox_pago_mode,
        fecha_prox_pago_manual=fecha_prox_pago_manual,
        limite=limite,
    )
    columnas = [
        "documento",
        "PrimerNombre",
        "PrimerApellido",
        "ValorCredito",
        "Saldo",
        "DiasCredito",
        "dia",
        "FechaProxPago",
        "Abono",
    ]
    indices = [headers.index(columna) for columna in columnas]
    preview = [[row[index] for index in indices] for row in rows]
    return columnas, preview


def generar_bitacora_cargue(
    normalizacion,
    *,
    archivo_cliente,
    archivo_csv,
    archivo_csv_listo,
    archivo_excel,
    mapeo,
    preview,
    header_row,
    data_start_row,
    responsable="",
    auditoria=None,
    advertencia_excel="",
):
    global ULTIMO_RESUMEN_EJECUTIVO
    auditoria = auditoria or _auditoria_desde_responsable(responsable)
    output_dir = REPORTS_DIR / "cargue_rutas"
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    path = output_dir / f"bitacora_cargue_{timestamp}.txt"
    resumen = resumen_validacion(normalizacion)
    errores = [
        f"- Fila {registro.fila_origen} / doc {registro.documento or 'sin documento'}: "
        + "; ".join(registro.errores)
        for registro in normalizacion.registros
        if registro.errores
    ][:50]
    contenido = (
        "BITACORA CARGUE DE RUTAS\n"
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        "Auditoria:\n"
        + _formatear_auditoria(auditoria)
        + "\n"
        + "Archivos:\n"
        f"- Cliente: {archivo_cliente}\n"
        f"- Excel respaldo: {archivo_excel}\n"
        f"- CSV interno: {archivo_csv}\n"
        f"- CSV LISTO: {archivo_csv_listo}\n\n"
        "Lectura:\n"
        f"- Hoja: {preview.hoja}\n"
        f"- Fila encabezados: {header_row}\n"
        f"- Fila inicial clientes: {data_start_row}\n"
        f"- Columnas detectadas: {len(preview.columnas)}\n\n"
        "Resumen:\n"
        f"- Total registros: {resumen['total']}\n"
        f"- Validos: {resumen['validos']}\n"
        f"- Errores bloqueantes: {resumen['errores_bloqueantes']}\n"
        f"- Advertencias: {resumen['advertencias']}\n"
        f"- Ajustes automaticos: {resumen['ajustes_automaticos']}\n"
        f"- Abonos informados: {resumen['abonos_informados']}\n"
        f"- Abonos reconstruidos: {resumen['abonos_reconstruidos']}\n\n"
        "Mapeo:\n"
        + "\n".join(_formatear_mapeo(mapeo))
        + "\n\nMensajes generales:\n"
        + ("\n".join(normalizacion.errores_generales) if normalizacion.errores_generales else "- Sin mensajes generales")
        + (f"\n{advertencia_excel.strip()}" if advertencia_excel.strip() else "")
        + "\n\nPrimeros errores:\n"
        + ("\n".join(errores) if errores else "- Sin errores bloqueantes")
        + "\n"
    )
    path.write_text(contenido, encoding="utf-8")
    central_path, central_error = registrar_bitacora_central(
        auditoria=auditoria,
        archivo_cliente=archivo_cliente,
        archivo_csv=archivo_csv,
        archivo_csv_listo=archivo_csv_listo,
        archivo_excel=archivo_excel,
        resumen=resumen,
        total_errores=len(errores),
        bitacora_local=path,
    )
    estado_onedrive = obtener_ultimo_estado_onedrive()
    resumen_ejecutivo, resumen_error = generar_resumen_ejecutivo_cargue(
        normalizacion,
        auditoria=auditoria,
        archivo_cliente=archivo_cliente,
        archivo_csv_listo=archivo_csv_listo,
        archivo_excel=archivo_excel,
        bitacora_local=path,
        estado_onedrive=estado_onedrive,
        resumen=resumen,
        output_dir=output_dir,
        timestamp=timestamp,
    )
    ULTIMO_RESUMEN_EJECUTIVO = str(resumen_ejecutivo or "")
    resumen_ejecutivo_texto = (
        f"- Generado en: {resumen_ejecutivo}\n"
        if resumen_ejecutivo
        else f"- No generado: {resumen_error}\n"
    )
    if central_path:
        path.write_text(
            contenido
            + f"\nBitacora central:\n- Registrada en: {central_path}\n"
            + f"- Estado OneDrive: {estado_onedrive}\n"
            + "\nResumen ejecutivo:\n"
            + resumen_ejecutivo_texto,
            encoding="utf-8",
        )
    else:
        path.write_text(
            contenido
            + "\nBitacora central:\n"
            + f"- No registrada: {central_error}\n"
            + f"- Estado OneDrive: {estado_onedrive}\n"
            + "\nResumen ejecutivo:\n"
            + resumen_ejecutivo_texto,
            encoding="utf-8",
        )
    return path


def generar_resumen_ejecutivo_cargue(
    normalizacion,
    *,
    auditoria,
    archivo_cliente,
    archivo_csv_listo,
    archivo_excel,
    bitacora_local,
    estado_onedrive,
    resumen,
    output_dir,
    timestamp,
):
    path = output_dir / f"resumen_ejecutivo_cargue_{timestamp}.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        return None, "Falta instalar openpyxl para generar el resumen ejecutivo."

    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Resumen ejecutivo"

        title_fill = PatternFill("solid", fgColor="0B1B26")
        subtitle_fill = PatternFill("solid", fgColor="102A3A")
        accent_fill = PatternFill("solid", fgColor="13A8D3")
        soft_fill = PatternFill("solid", fgColor="EAF6FA")
        ok_fill = PatternFill("solid", fgColor="DDF5E7")
        warn_fill = PatternFill("solid", fgColor="FFF2CC")
        error_fill = PatternFill("solid", fgColor="FCE4D6")
        thin = Side(style="thin", color="D7E8EF")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        sheet.merge_cells("A1:D1")
        sheet["A1"] = "Resumen ejecutivo de cargue de rutas"
        sheet["A1"].fill = title_fill
        sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 30

        sheet.merge_cells("A2:D2")
        sheet["A2"] = "Control generado automaticamente por Procesos AM - Apostamos por la tecnologia"
        sheet["A2"].fill = subtitle_fill
        sheet["A2"].font = Font(color="D9F5FF", size=10)
        sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

        row = 4
        row = _write_section(sheet, row, "Auditoria", accent_fill, border)
        auditoria_rows = [
            ("Fecha", datetime.now().strftime("%Y-%m-%d")),
            ("Hora", datetime.now().strftime("%H:%M:%S")),
            ("Nombre de BD", auditoria.get("nombre_bd") or "No informado"),
            ("Tipo de cargue", auditoria.get("tipo_cargue") or "Produccion"),
            ("Responsable", auditoria.get("responsable_nombre") or "No informado"),
            ("Correo", auditoria.get("responsable_correo") or "No informado"),
            ("Hostname", auditoria.get("hostname") or "No detectado"),
            ("IP equipo", auditoria.get("ip") or "No detectada"),
            ("Usuario Windows", auditoria.get("usuario_windows") or "No detectado"),
        ]
        row = _write_pairs(sheet, row, auditoria_rows, soft_fill, border)

        row += 1
        row = _write_section(sheet, row, "Archivos generados", accent_fill, border)
        archivo_rows = [
            ("Archivo cliente", str(archivo_cliente)),
            ("CSV final para LISTO", str(archivo_csv_listo)),
            ("Excel respaldo", str(archivo_excel)),
            ("Bitacora tecnica", str(bitacora_local)),
        ]
        row = _write_pairs(sheet, row, archivo_rows, soft_fill, border)

        row += 1
        row = _write_section(sheet, row, "Resumen de validacion", accent_fill, border)
        indicadores = [
            ("Total registros", resumen["total"]),
            ("Registros validos", resumen["validos"]),
            ("Errores bloqueantes", resumen["errores_bloqueantes"]),
            ("Advertencias", resumen["advertencias"]),
            ("Ajustes automaticos", resumen["ajustes_automaticos"]),
            ("Abonos informados", resumen["abonos_informados"]),
            ("Abonos reconstruidos", resumen["abonos_reconstruidos"]),
        ]
        for label, value in indicadores:
            fill = ok_fill
            if label == "Errores bloqueantes" and value:
                fill = error_fill
            elif label in {"Advertencias", "Ajustes automaticos", "Abonos reconstruidos"} and value:
                fill = warn_fill
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=value)
            sheet.cell(row=row, column=1).fill = soft_fill
            sheet.cell(row=row, column=2).fill = fill
            for column in range(1, 3):
                cell = sheet.cell(row=row, column=column)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1

        row += 1
        row = _write_section(sheet, row, "Estado OneDrive / SharePoint", accent_fill, border)
        row = _write_pairs(sheet, row, [("Estado", estado_onedrive)], soft_fill, border)

        row += 1
        row = _write_section(sheet, row, "Novedades para revisar", accent_fill, border)
        novedades = _novedades_resumen(normalizacion)
        if not novedades:
            novedades = ["Sin errores bloqueantes ni advertencias generales."]
        for item in novedades[:25]:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = sheet.cell(row=row, column=1, value=item)
            cell.fill = soft_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1

        widths = {"A": 24, "B": 34, "C": 24, "D": 54}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.font = cell.font.copy(name="Segoe UI")

        workbook.save(path)
        return path, None
    except Exception as error:
        return None, str(error)


def _write_section(sheet, row, title, fill, border):
    from openpyxl.styles import Alignment, Font

    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = fill
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border
    sheet.row_dimensions[row].height = 24
    return row + 1


def _write_pairs(sheet, row, pairs, fill, border):
    from openpyxl.styles import Alignment, Font

    for label, value in pairs:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        for column in range(1, 5):
            cell = sheet.cell(row=row, column=column)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
    return row


def _novedades_resumen(normalizacion):
    novedades = []
    for mensaje in normalizacion.errores_generales:
        novedades.append(f"General: {mensaje}")
    for registro in normalizacion.registros:
        if registro.errores:
            novedades.append(
                f"Fila {registro.fila_origen} / doc {registro.documento or 'sin documento'}: "
                + "; ".join(registro.errores)
            )
    return novedades


def obtener_ultimo_resumen_ejecutivo():
    return ULTIMO_RESUMEN_EJECUTIVO or ""


def procesar_archivo(path):
    archivo = validar_archivo(path)
    datos = f"Archivo recibido: {archivo}"
    datos = limpiar_datos(datos)
    resultado = calcular(datos)
    salida = REPORTS_DIR / "cargue_rutas_resultado.txt"
    return exportar_resultado(resultado, salida)


def registrar_bitacora_central(
    *,
    auditoria,
    archivo_cliente,
    archivo_csv,
    archivo_csv_listo,
    archivo_excel,
    resumen,
    total_errores,
    bitacora_local,
):
    global ULTIMO_ESTADO_ONEDRIVE
    carpeta = _buscar_carpeta_auditoria_central()
    if carpeta is None:
        ULTIMO_ESTADO_ONEDRIVE = (
            "No se encontro la carpeta sincronizada. No se pudo registrar en SharePoint/OneDrive. "
            "Revise OneDrive o sincronice la biblioteca."
        )
        return None, "No se encontro la carpeta sincronizada de OneDrive/SharePoint."

    onedrive_activo = _onedrive_en_ejecucion()
    if not onedrive_activo:
        ULTIMO_ESTADO_ONEDRIVE = (
            "Se encontro la carpeta local, pero OneDrive no esta en ejecucion. "
            "El registro quedara local hasta que abra o reanude OneDrive."
        )

    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return None, "Falta instalar openpyxl para crear la bitacora central en Excel."

    try:
        carpeta.mkdir(parents=True, exist_ok=True)
        destino = carpeta / AUDITORIA_CENTRAL_FILE
        existe = destino.exists()
        campos = _campos_bitacora_central()
        ahora = datetime.now()
        fila = {
            "id_cargue": _id_cargue(auditoria, archivo_cliente, ahora),
            "fecha": ahora.strftime("%Y-%m-%d"),
            "hora": ahora.strftime("%H:%M:%S"),
            "nombre_bd": auditoria.get("nombre_bd") or "No informado",
            "tipo_cargue": auditoria.get("tipo_cargue") or "Produccion",
            "responsable": auditoria.get("responsable_nombre") or "No informado",
            "correo": auditoria.get("responsable_correo") or "No informado",
            "hostname": auditoria.get("hostname") or "No detectado",
            "ip_equipo": auditoria.get("ip") or "No detectada",
            "usuario_windows": auditoria.get("usuario_windows") or "No detectado",
            "archivo_cliente": str(archivo_cliente),
            "archivo_csv_listo": str(archivo_csv_listo or archivo_csv),
            "archivo_excel_respaldo": str(archivo_excel),
            "bitacora_local": str(bitacora_local),
            "total_registros": resumen["total"],
            "registros_validos": resumen["validos"],
            "errores_bloqueantes": resumen["errores_bloqueantes"],
            "advertencias": resumen["advertencias"],
            "ajustes_automaticos": resumen["ajustes_automaticos"],
            "abonos_informados": resumen["abonos_informados"],
            "abonos_reconstruidos": resumen["abonos_reconstruidos"],
        }

        if existe:
            workbook = load_workbook(destino)
            sheet = _obtener_hoja_cargues(workbook, campos)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Cargues"
            _preparar_hoja_bitacora(sheet, campos)
            _preparar_hoja_procesos_am(workbook)

        fila_excel = _buscar_fila_por_id(sheet, fila["id_cargue"])
        valores = [fila.get(campo, "") for campo in campos]
        if fila_excel:
            for column, value in enumerate(valores, start=1):
                sheet.cell(row=fila_excel, column=column, value=value)
        else:
            sheet.append(valores)

        _aplicar_estilo_bitacora(sheet, campos, Table, TableStyleInfo, Font, PatternFill, Alignment, Border, Side)
        _aplicar_estilo_procesos_am(
            _preparar_hoja_procesos_am(workbook),
            Table,
            TableStyleInfo,
            Font,
            PatternFill,
            Alignment,
            Border,
            Side,
        )
        workbook.save(destino)
        if onedrive_activo:
            ULTIMO_ESTADO_ONEDRIVE = (
                f"Registro escrito en carpeta sincronizada a las {datetime.now().strftime('%H:%M:%S')}. "
                "OneDrive debe subirlo a SharePoint. Para confirmar subida al servidor se requiere API de SharePoint/Graph."
            )
        return destino, None
    except Exception as error:
        ULTIMO_ESTADO_ONEDRIVE = (
            "No se pudo escribir la bitacora central. Dele play/reanude OneDrive o revise la sincronizacion."
        )
        return None, str(error)


def _buscar_carpeta_auditoria_central():
    ruta_onedrive = Path.home() / ONEDRIVE_EMPRESA
    if not ruta_onedrive.is_dir():
        return None

    for ruta_relativa in RUTAS_AUDITORIA_ONEDRIVE:
        candidata = ruta_onedrive / ruta_relativa
        if candidata.is_dir():
            return candidata / AUDITORIA_CENTRAL_DIR
    return None


def obtener_ultimo_estado_onedrive():
    return ULTIMO_ESTADO_ONEDRIVE or "Estado OneDrive no evaluado."


def _onedrive_en_ejecucion():
    try:
        result = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq OneDrive.exe"],
            capture_output=True,
            text=True,
            timeout=3,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except Exception:
        return False
    return "OneDrive.exe" in result.stdout


def eliminar_cargue_prueba_central(*, auditoria, archivo_cliente):
    id_cargue = _id_cargue(auditoria, archivo_cliente, datetime.now())
    return eliminar_cargue_prueba_por_id(id_cargue)


def listar_cargues_prueba_central():
    carpeta = _buscar_carpeta_auditoria_central()
    if carpeta is None:
        return [], "No se encontro la carpeta sincronizada de OneDrive/SharePoint."

    destino = carpeta / AUDITORIA_CENTRAL_FILE
    if not destino.exists():
        return [], "No existe bitacora central para listar pruebas."

    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "Falta instalar openpyxl para leer la bitacora central."

    try:
        campos = _campos_bitacora_central()
        workbook = load_workbook(destino, data_only=True)
        sheet = _obtener_hoja_cargues(workbook, campos)
        headers = _leer_headers_bitacora(sheet)
        registros = []
        for row in range(5, sheet.max_row + 1):
            item = {
                header: sheet.cell(row=row, column=column).value
                for column, header in enumerate(headers, start=1)
                if header
            }
            tipo = str(item.get("tipo_cargue") or "").strip().lower()
            if tipo != "prueba":
                continue
            item["_fila_excel"] = row
            item["id_cargue"] = str(item.get("id_cargue") or "")
            registros.append(item)
        registros.sort(
            key=lambda item: f"{item.get('fecha') or ''} {item.get('hora') or ''}",
            reverse=True,
        )
        return registros, None
    except Exception as error:
        return [], f"No se pudieron listar los cargues de prueba: {error}"


def eliminar_cargue_prueba_por_id(id_cargue):
    carpeta = _buscar_carpeta_auditoria_central()
    if carpeta is None:
        return False, "No se encontro la carpeta sincronizada de OneDrive/SharePoint."

    destino = carpeta / AUDITORIA_CENTRAL_FILE
    if not destino.exists():
        return False, "No existe bitacora central para eliminar."

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return False, "Falta instalar openpyxl para modificar la bitacora central."

    try:
        campos = _campos_bitacora_central()
        workbook = load_workbook(destino)
        sheet = _obtener_hoja_cargues(workbook, campos)
        fila_excel = _buscar_fila_por_id(sheet, id_cargue)
        if not fila_excel:
            return False, "No se encontro el registro de prueba seleccionado."
        headers = _leer_headers_bitacora(sheet)
        tipo_col = headers.index("tipo_cargue") + 1 if "tipo_cargue" in headers else 5
        tipo = str(sheet.cell(row=fila_excel, column=tipo_col).value or "").strip().lower()
        if tipo != "prueba":
            return False, "El registro encontrado no esta marcado como Prueba."
        sheet.delete_rows(fila_excel, 1)
        _aplicar_estilo_bitacora(sheet, campos, Table, TableStyleInfo, Font, PatternFill, Alignment, Border, Side)
        _aplicar_estilo_procesos_am(
            _preparar_hoja_procesos_am(workbook),
            Table,
            TableStyleInfo,
            Font,
            PatternFill,
            Alignment,
            Border,
            Side,
        )
        workbook.save(destino)
        return True, "Registro de prueba eliminado de la bitacora central."
    except Exception as error:
        return False, f"No se pudo eliminar el registro de prueba: {error}"


def eliminar_cargues_prueba_por_ids(ids_cargue):
    ids = {str(id_cargue or "").strip() for id_cargue in ids_cargue if str(id_cargue or "").strip()}
    if not ids:
        return False, "No hay cargues de prueba seleccionados para eliminar."

    carpeta = _buscar_carpeta_auditoria_central()
    if carpeta is None:
        return False, "No se encontro la carpeta sincronizada de OneDrive/SharePoint."

    destino = carpeta / AUDITORIA_CENTRAL_FILE
    if not destino.exists():
        return False, "No existe bitacora central para eliminar."

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return False, "Falta instalar openpyxl para modificar la bitacora central."

    try:
        campos = _campos_bitacora_central()
        workbook = load_workbook(destino)
        sheet = _obtener_hoja_cargues(workbook, campos)
        headers = _leer_headers_bitacora(sheet)
        tipo_col = headers.index("tipo_cargue") + 1 if "tipo_cargue" in headers else 5
        eliminados = 0

        for row in range(sheet.max_row, 4, -1):
            id_actual = str(sheet.cell(row=row, column=1).value or "").strip()
            tipo = str(sheet.cell(row=row, column=tipo_col).value or "").strip().lower()
            if id_actual in ids and tipo == "prueba":
                sheet.delete_rows(row, 1)
                eliminados += 1

        if eliminados == 0:
            return False, "No se encontraron registros de prueba para eliminar."

        _aplicar_estilo_bitacora(sheet, campos, Table, TableStyleInfo, Font, PatternFill, Alignment, Border, Side)
        _aplicar_estilo_procesos_am(
            _preparar_hoja_procesos_am(workbook),
            Table,
            TableStyleInfo,
            Font,
            PatternFill,
            Alignment,
            Border,
            Side,
        )
        workbook.save(destino)
        return True, f"Se eliminaron {eliminados} registros de prueba de la bitacora central."
    except Exception as error:
        return False, f"No se pudieron eliminar los registros de prueba: {error}"


def _campos_bitacora_central():
    return [
        "id_cargue",
        "fecha",
        "hora",
        "nombre_bd",
        "tipo_cargue",
        "responsable",
        "correo",
        "hostname",
        "ip_equipo",
        "usuario_windows",
        "archivo_cliente",
        "archivo_csv_listo",
        "archivo_excel_respaldo",
        "bitacora_local",
        "total_registros",
        "registros_validos",
        "errores_bloqueantes",
        "advertencias",
        "ajustes_automaticos",
        "abonos_informados",
        "abonos_reconstruidos",
    ]


def _id_cargue(auditoria, archivo_cliente, fecha_hora):
    responsable = auditoria.get("responsable_nombre") or "sin_responsable"
    hostname = auditoria.get("hostname") or "sin_hostname"
    nombre_bd = auditoria.get("nombre_bd") or "sin_bd"
    tipo_cargue = auditoria.get("tipo_cargue") or "Produccion"
    return "|".join(
        [
            fecha_hora.strftime("%Y-%m-%d"),
            str(archivo_cliente).lower(),
            nombre_bd.lower(),
            tipo_cargue.lower(),
            responsable.lower(),
            hostname.lower(),
        ]
    )


def _obtener_hoja_cargues(workbook, campos):
    if "Cargues" in workbook.sheetnames:
        sheet = workbook["Cargues"]
        if sheet.max_row < 4:
            _preparar_hoja_bitacora(sheet, campos)
        if "Procesos AM" not in workbook.sheetnames:
            _preparar_hoja_procesos_am(workbook)
        return sheet
    sheet = workbook.create_sheet("Cargues")
    _preparar_hoja_bitacora(sheet, campos)
    if "Procesos AM" not in workbook.sheetnames:
        _preparar_hoja_procesos_am(workbook)
    return sheet


def _preparar_hoja_bitacora(sheet, campos):
    sheet["A1"] = "Bitacora central de cargues"
    sheet["A2"] = "Control generado automaticamente por Procesos AM."
    sheet.append([])
    sheet.append(campos)
    sheet.freeze_panes = "A5"


def _preparar_hoja_procesos_am(workbook):
    if "Procesos AM" in workbook.sheetnames:
        sheet = workbook["Procesos AM"]
    else:
        sheet = workbook.create_sheet("Procesos AM", 0)
    sheet["A1"] = "Bitacora central de Procesos AM"
    sheet["A2"] = "Reservado para el historico del monitoreo 7 AM."
    for column, campo in enumerate(["fecha", "hora", "responsable", "hostname", "revision", "estado"], start=1):
        sheet.cell(row=4, column=column, value=campo)
    return sheet


def _buscar_fila_por_id(sheet, id_cargue):
    for row in range(5, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == id_cargue:
            return row
    return None


def _leer_headers_bitacora(sheet):
    return [
        str(sheet.cell(row=4, column=column).value or "").strip()
        for column in range(1, sheet.max_column + 1)
    ]


def _aplicar_estilo_bitacora(sheet, campos, Table, TableStyleInfo, Font, PatternFill, Alignment, Border, Side):
    max_col = len(campos)
    max_row = max(sheet.max_row, 4)
    title_fill = PatternFill("solid", fgColor="0B1B26")
    subtitle_fill = PatternFill("solid", fgColor="102A3A")
    header_fill = PatternFill("solid", fgColor="13A8D3")
    header_font = Font(color="FFFFFF", bold=True)
    text_font = Font(color="0B1B26")
    muted_font = Font(color="4B6475")
    thin = Side(style="thin", color="D7E8EF")

    for column, campo in enumerate(campos, start=1):
        sheet.cell(row=4, column=column, value=campo)

    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row in {1, 2}:
            sheet.unmerge_cells(str(merged))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    sheet["A1"].fill = title_fill
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"].fill = subtitle_fill
    sheet["A2"].font = Font(color="D9F5FF", size=10)
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 22

    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet.row_dimensions[4].height = 32

    for row in sheet.iter_rows(min_row=5, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = text_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    widths = {
        "A": 18,
        "B": 12,
        "C": 11,
        "D": 24,
        "E": 16,
        "F": 22,
        "G": 28,
        "H": 22,
        "I": 16,
        "J": 24,
        "K": 46,
        "L": 34,
        "M": 34,
        "N": 42,
        "O": 14,
        "P": 14,
        "Q": 18,
        "R": 14,
        "S": 18,
        "T": 18,
        "U": 20,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["A"].hidden = True

    for cell in sheet["A"]:
        cell.font = muted_font

    table_ref = f"A4:U{max_row}"
    if max_row > 4:
        if "TablaBitacoraCargues" not in sheet.tables:
            table = Table(displayName="TablaBitacoraCargues", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        else:
            sheet.tables["TablaBitacoraCargues"].ref = table_ref
        sheet.auto_filter.ref = table_ref
    else:
        if "TablaBitacoraCargues" in sheet.tables:
            del sheet.tables["TablaBitacoraCargues"]
        sheet.auto_filter.ref = "A4:U4"


def _aplicar_estilo_procesos_am(sheet, Table, TableStyleInfo, Font, PatternFill, Alignment, Border, Side):
    max_row = max(sheet.max_row, 4)
    title_fill = PatternFill("solid", fgColor="0B1B26")
    subtitle_fill = PatternFill("solid", fgColor="102A3A")
    header_fill = PatternFill("solid", fgColor="13A8D3")
    header_font = Font(color="FFFFFF", bold=True)
    text_font = Font(color="0B1B26")
    thin = Side(style="thin", color="D7E8EF")

    for column, campo in enumerate(["fecha", "hora", "responsable", "hostname", "revision", "estado"], start=1):
        sheet.cell(row=4, column=column, value=campo)

    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row in {1, 2}:
            sheet.unmerge_cells(str(merged))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    sheet["A1"].fill = title_fill
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"].fill = subtitle_fill
    sheet["A2"].font = Font(color="D9F5FF", size=10)
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 22
    sheet.row_dimensions[4].height = 32

    for cell in sheet[4][:6]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in sheet.iter_rows(min_row=5, max_row=max_row, max_col=6):
        for cell in row:
            cell.font = text_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    widths = {
        "A": 12,
        "B": 11,
        "C": 24,
        "D": 22,
        "E": 18,
        "F": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    table_ref = f"A4:F{max_row}"
    if max_row > 4:
        if "TablaProcesosAM" not in sheet.tables:
            table = Table(displayName="TablaProcesosAM", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        else:
            sheet.tables["TablaProcesosAM"].ref = table_ref
        sheet.auto_filter.ref = table_ref
    else:
        if "TablaProcesosAM" in sheet.tables:
            del sheet.tables["TablaProcesosAM"]
        sheet.auto_filter.ref = "A4:F4"


def _auditoria_desde_responsable(responsable):
    return {
        "responsable_nombre": responsable or "",
        "responsable_correo": "",
        "nombre_bd": "",
        "tipo_cargue": "Produccion",
        "hostname": "",
        "ip": "",
        "usuario_windows": "",
    }


def _formatear_auditoria(auditoria):
    auditoria = auditoria or {}
    return (
        f"- Nombre de BD: {auditoria.get('nombre_bd') or 'No informado'}\n"
        f"- Tipo de cargue: {auditoria.get('tipo_cargue') or 'Produccion'}\n"
        f"- Responsable: {auditoria.get('responsable_nombre') or 'No informado'}\n"
        f"- Correo: {auditoria.get('responsable_correo') or 'No informado'}\n"
        f"- Hostname: {auditoria.get('hostname') or 'No detectado'}\n"
        f"- IP equipo: {auditoria.get('ip') or 'No detectada'}\n"
        f"- Usuario Windows: {auditoria.get('usuario_windows') or 'No detectado'}\n"
    )


def _formatear_mapeo(mapeo):
    if isinstance(mapeo, dict):
        return [f"- {campo}: {columna or 'Pendiente'}" for campo, columna in mapeo.items()]

    return [
        f"- {item.campo.nombre}: {item.columna_origen or 'Pendiente'} ({item.confianza}%)"
        for item in mapeo
    ]


def _resumen_periodicidad(registros):
    nombres = {
        1: "1 - Diario",
        2: "2 - Semanal",
        3: "3 - Quincenal",
        4: "4 - Mensual",
        None: "Sin codigo",
    }
    conteo = {1: 0, 2: 0, 3: 0, 4: 0, None: 0}
    for registro in registros:
        codigo = registro.periodicidad_codigo if registro.periodicidad_codigo in {1, 2, 3, 4} else None
        conteo[codigo] += 1
    return [f"- {nombres[codigo]}: {conteo[codigo]}" for codigo in (1, 2, 3, 4, None)]


def _resumen_abonos(registros):
    estados = {
        "informado": "Abono informado por cliente",
        "requiere_reconstruccion": "Pendiente de reconstruir",
        "reconstruido": "Abono reconstruido por sistema",
        "sin_columna": "Sin columna de abono",
    }
    conteo = {estado: 0 for estado in estados}
    conteo["otro"] = 0
    for registro in registros:
        if registro.abono_estado in conteo:
            conteo[registro.abono_estado] += 1
        else:
            conteo["otro"] += 1
    lineas = [f"- {label}: {conteo[estado]}" for estado, label in estados.items()]
    if conteo["otro"]:
        lineas.append(f"- Otro: {conteo['otro']}")
    return lineas
